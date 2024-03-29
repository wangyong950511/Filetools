import os
import openpyxl

# 设置工作路径
# 获取当前工作目录
current_directory = os.getcwd()
print("当前工作目录为：", current_directory)
# 设置新的工作目录
new_directory = "/Users/wangyong/Documents/CASH/CASHrawdata/Cell.nosync/YHRS-2024.3.15"
os.chdir(new_directory)
# 获取当前工作目录
current_directory = os.getcwd()
print("当前工作目录为：", current_directory)

name_path = r'/Users/wangyong/Documents/CASH/CASHrawdata/分析/YHRS-2024.3.15/exchange模版.xlsx'
# 打开工作簿
wb = openpyxl.load_workbook(name_path)
# 获取数据源表格 0-第一列，1-第二列
sht = wb.worksheets[1]
# 获取 A 列和 B 列的最大行数
max_row = max(sht.max_row, sht.max_column)
# 遍历 A 列和 B 列，将非空数据添加到字典中
for row in range(1, max_row + 1):
    key = sht.cell(row=row, column=1).value
    value = sht.cell(row=row, column=2).value
    if key is not None and value is not None:
        file_name = os.path.abspath(key)
        file_rename = os.path.abspath(value)
        os.rename(file_name, file_rename)
